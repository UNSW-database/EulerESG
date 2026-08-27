"""
Excel Export Module for ESG Analysis Results

This module handles exporting ESG compliance analysis results to Excel format,
matching the structure of the MCG-Financials template.
"""

import pandas as pd
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional, Any
from loguru import logger
import json

from .file_manager import file_manager


class ExcelExporter:
    """Export ESG analysis results to Excel format"""
    
    def __init__(self, output_dir: Optional[Path] = None):
        """
        Initialize Excel Exporter
        
        Args:
            output_dir: Directory to save Excel files. Defaults to 'outputs/excel'
        """
        if output_dir is None:
            # Canonical location: uploads/outputs/excel
            self.output_dir = Path(file_manager.outputs_dir) / "excel"
        else:
            self.output_dir = Path(output_dir)
        
        # Ensure output directory exists
        self.output_dir.mkdir(parents=True, exist_ok=True)
        logger.info(f"Excel exporter initialized with output directory: {self.output_dir}")
    
    def export_analysis_results(
        self,
        metric_analyses: List[Dict[str, Any]],
        industry: str,
        semi_industry: str,
        company_name: Optional[str] = None,
        report_id: Optional[str] = None
    ) -> Path:
        """
        Export analysis results to Excel file
        
        Args:
            metric_analyses: List of analyzed metrics with results
            industry: Main industry category
            semi_industry: Sub-industry for specific metrics
            company_name: Name of the company being analyzed
            report_id: Unique identifier for the report
            
        Returns:
            Path: Path to the generated Excel file
        """
        try:
            def _status_of(metric: Dict[str, Any]) -> str:
                raw = (
                    metric.get("Disclosure Status")
                    or metric.get("disclosure_status")
                    or metric.get("Model Disclosure Status")
                    or ""
                )
                normalized = str(raw).strip().lower().replace("-", "_").replace(" ", "_")
                if "not_clear" in normalized or "unclear" in normalized:
                    return "partially_disclosed"
                if "not" in normalized:
                    return "not_disclosed"
                if "partial" in normalized:
                    return "partially_disclosed"
                if "full" in normalized or normalized == "disclosed":
                    return "fully_disclosed"
                return normalized

            def _status_label(metric: Dict[str, Any]) -> str:
                status = _status_of(metric)
                return {
                    "fully_disclosed": "Disclosed",
                    "partially_disclosed": "Partially Disclosed",
                    "not_disclosed": "Not Disclosed",
                }.get(status, status.replace("_", " ").title())

            # Prepare data for Excel
            excel_data = []
            
            for metric in metric_analyses:
                row = {
                    "Metric": metric.get("Metric", metric.get("metric_name", "")),
                    "Category": metric.get("Category", metric.get("category", "")),
                    "Unit": metric.get("Unit", metric.get("unit", "")),
                    "Code": metric.get("Code", metric.get("metric_code", metric.get("metric_id", ""))),
                    "Topic": metric.get("Topic", metric.get("topic", "")),
                    "Type": metric.get("Type", metric.get("type", "")),
                    "Definition": metric.get("Definition", metric.get("definition", "")),
                    "Value": self._format_value(metric.get("Value", metric.get("value"))),
                    "Selected Year": metric.get("Selected Year", metric.get("selected_year")),
                    "Annual Values": json.dumps(
                        metric.get("Year Values", metric.get("year_values", [])) or [],
                        ensure_ascii=False,
                    ),
                    "Page": self._format_page(metric.get("Page", metric.get("page"))),
                    "Context": metric.get("Context", metric.get("context", "")),
                    "Disclosure Status": _status_label(metric),
                    "LLM Analysis": metric.get("LLM Analysis", metric.get("reasoning", ""))
                }
                excel_data.append(row)
            
            # Create DataFrame
            df = pd.DataFrame(excel_data)
            
            # Generate filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            safe_company = self._sanitize_filename(company_name) if company_name else "Company"
            safe_industry = self._sanitize_filename(semi_industry)
            filename = f"{safe_company}_{safe_industry}_{timestamp}.xlsx"
            filepath = self.output_dir / filename
            
            # Create Excel writer with formatting
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                # Write main data
                sheet_name = self._truncate_sheet_name(semi_industry)
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Get the worksheet
                worksheet = writer.sheets[sheet_name]
                
                # Apply formatting
                self._apply_excel_formatting(worksheet, df)
                
                # Add metadata sheet
                metadata = {
                    "Analysis Date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "Company": company_name or "Unknown",
                    "Industry": industry,
                    "Sub-Industry": semi_industry,
                    "Report ID": report_id or "N/A",
                    "Total Metrics": len(metric_analyses),
                    "Disclosed": sum(1 for m in metric_analyses if _status_of(m) == "fully_disclosed"),
                    "Partially Disclosed": sum(1 for m in metric_analyses if _status_of(m) == "partially_disclosed"),
                    "Not Disclosed": sum(1 for m in metric_analyses if _status_of(m) == "not_disclosed")
                }
                
                metadata_df = pd.DataFrame(list(metadata.items()), columns=["Field", "Value"])
                metadata_df.to_excel(writer, sheet_name="Summary", index=False)
                
                # Format metadata sheet
                metadata_worksheet = writer.sheets["Summary"]
                self._apply_metadata_formatting(metadata_worksheet, metadata_df)
            
            logger.info(f"Excel file successfully created: {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"Error creating Excel file: {str(e)}")
            raise
    
    def _format_value(self, value: Any) -> Any:
        """Format value field for Excel"""
        if value is None or value == "null":
            return None
        if value == "not specific":
            return "Not Specific"
        return value
    
    def _format_page(self, page: Any) -> str:
        """Format page field for Excel"""
        if page is None or page == "null":
            return ""
        if isinstance(page, (list, tuple)):
            return ", ".join(str(p) for p in page)
        return str(page)
    
    def _sanitize_filename(self, name: str) -> str:
        """Sanitize string for use in filename"""
        if not name:
            return "Unknown"
        # Remove/replace invalid filename characters
        invalid_chars = '<>:"/\\|?*'
        for char in invalid_chars:
            name = name.replace(char, '_')
        # Limit length
        return name[:50]
    
    def _truncate_sheet_name(self, name: str) -> str:
        """Truncate sheet name to Excel's 31 character limit"""
        if len(name) > 31:
            return name[:28] + "..."
        return name
    
    def _apply_excel_formatting(self, worksheet, df):
        """Apply formatting to the main data worksheet"""
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        
        # Header formatting
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        for col in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        # Limit max width to 50 characters
                        max_length = max(max_length, min(len(str(cell.value)), 50))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Apply text wrapping to verbose text columns using dynamic indexes
        wrap_alignment = Alignment(wrap_text=True, vertical="top")
        column_index = {name: idx + 1 for idx, name in enumerate(df.columns)}
        for row in range(2, len(df) + 2):
            for col_name in ("Definition", "Context", "LLM Analysis"):
                col_idx = column_index.get(col_name)
                if col_idx:
                    worksheet.cell(row=row, column=col_idx).alignment = wrap_alignment
        
        # Color code disclosure status
        status_colors = {
            "Disclosed": "C6EFCE",  # Light green
            "Partially Disclosed": "FFEB9C",  # Light yellow
            "Not Disclosed": "FFC7CE"  # Light red
        }
        
        status_col = column_index.get("Disclosure Status")
        if status_col:
            for row in range(2, len(df) + 2):
                cell = worksheet.cell(row=row, column=status_col)
                if cell.value in status_colors:
                    cell.fill = PatternFill(start_color=status_colors[cell.value], 
                                           end_color=status_colors[cell.value], 
                                           fill_type="solid")
    
    def _apply_metadata_formatting(self, worksheet, df):
        """Apply formatting to the metadata worksheet"""
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        
        # Header formatting
        header_font = Font(bold=True)
        for col in range(1, 3):
            worksheet.cell(row=1, column=col).font = header_font
        
        # Auto-adjust column widths
        worksheet.column_dimensions['A'].width = 20
        worksheet.column_dimensions['B'].width = 40
        
        # Bold the field names
        field_font = Font(bold=True)
        for row in range(2, len(df) + 2):
            worksheet.cell(row=row, column=1).font = field_font
    
    def export_template(self, semi_industry: str) -> Path:
        """
        Export a blank template for a specific industry
        
        Args:
            semi_industry: Sub-industry name
            
        Returns:
            Path: Path to the generated template file
        """
        try:
            # Load metrics for the industry
            from .metric_processor import MetricProcessor
            from .models import ProcessingConfig
            
            config = ProcessingConfig()
            processor = MetricProcessor(config)
            metrics_collection = processor.load_sasb_metrics_by_industry(semi_industry)
            
            # Prepare template data
            template_data = []
            for metric in metrics_collection.metrics:
                row = {
                    "Metric": metric.metric_name,
                    "Category": metric.sasb_category if hasattr(metric, 'sasb_category') else "",
                    "Unit": metric.unit or "",
                    "Code": metric.metric_code,
                    "Topic": metric.sasb_topic if hasattr(metric, 'sasb_topic') else "",
                    "Type": metric.sasb_type if hasattr(metric, 'sasb_type') else "",
                    "Value": None,
                    "Page": None,
                    "Context": None
                }
                template_data.append(row)
            
            # Create DataFrame
            df = pd.DataFrame(template_data)
            
            # Generate filename
            safe_industry = self._sanitize_filename(semi_industry)
            filename = f"Template_{safe_industry}.xlsx"
            filepath = self.output_dir / "templates" / filename
            filepath.parent.mkdir(parents=True, exist_ok=True)
            
            # Write to Excel
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                sheet_name = self._truncate_sheet_name(semi_industry)
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Apply formatting
                worksheet = writer.sheets[sheet_name]
                self._apply_excel_formatting(worksheet, df)
            
            logger.info(f"Template Excel file created: {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"Error creating template: {str(e)}")
            raise
